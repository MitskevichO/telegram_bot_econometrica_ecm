from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

def build_sample_file_bytes() -> bytes:
    df = pd.DataFrame(
        {
            "date": pd.date_range("2015-01-01", periods=60, freq="MS"),
            "Y": [100 + i * 0.8 for i in range(60)],
            "X": [50 + i * 0.5 for i in range(60)],
        }
    )
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="data")
    return bio.getvalue()
""" Excel-отчёт с 4 листами:
    1) Исходные данные
    2) История действий
    3) ECM модель (последняя построенная)
    4) Графики (с встроенными изображениями) """
def build_full_excel_report_bytes(df: pd.DataFrame, session: dict, images: dict[str, bytes]) -> bytes:

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        # 1. Исходные данные
        df.to_excel(writer, index=False, sheet_name="Исходные данные")
        # 2. История действий (если есть)
        history = session.get("history", [])
        if history:
            history_df = pd.DataFrame(history)
            history_df.to_excel(writer, index=False, sheet_name="История действий")
        else:
            # пустой лист с пояснением
            pd.DataFrame({"Информация": ["История действий отсутствует"]}).to_excel(
                writer, index=False, sheet_name="История действий"
            )
        # 3. Последняя ECM модель (только если есть результат)
        ecm_text = session.get("ecm_result")
        if ecm_text:
            # Создаём DataFrame с описанием модели (одна строка)
            ecm_df = pd.DataFrame({"ECM модель": [ecm_text]})
            ecm_df.to_excel(writer, index=False, sheet_name="ECM модель")
        else:
            pd.DataFrame({"Сообщение": ["ECM модель ещё не построена"]}).to_excel(
                writer, index=False, sheet_name="ECM модель"
            )
        # 4. Лист для графиков (пустая таблица, встроим картинки позже)
        pd.DataFrame({"График": []}).to_excel(writer, index=False, sheet_name="Графики")
    # Встраивание изображений на лист "Графики"
    if images:
        try:
            bio.seek(0)
            wb = load_workbook(bio)
            if "Графики" in wb.sheetnames:
                ws = wb["Графики"]
                row = 1
                for title, png in images.items():
                    img_bytes = io.BytesIO(png)
                    xl_img = XLImage(img_bytes)
                    xl_img.width = 600
                    xl_img.height = 400
                    # Записываем название графика в ячейку столбца A
                    ws.cell(row=row, column=1, value=title)
                    # Вставляем картинку справа от названия (столбец B)
                    ws.add_image(xl_img, f"B{row+1}")
                    row += 30   # отступ между графиками
                out = io.BytesIO()
                wb.save(out)
                return out.getvalue()
        except Exception:
            # Если встраивание не удалось, возвращаем Excel без картинок
            return bio.getvalue()
    else:
        return bio.getvalue()
